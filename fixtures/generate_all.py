#!/usr/bin/env python3
"""
Generate test fixtures for all supported formats.

This script creates test files with known content for testing
the Zig format implementations.

Formats generated:
- Arrow IPC (.arrow)
- Avro (.avro)
- ORC (.orc)
- Excel (.xlsx)
- Delta Lake (directory)
- Iceberg (directory)

Usage:
    cd fixtures
    pip install pyarrow fastavro pyorc openpyxl deltalake pyiceberg
    python generate_all.py
"""

import os
import json
import shutil
from pathlib import Path

import pyarrow as pa
import pyarrow.ipc as ipc
import pyarrow.parquet as pq

# Output directory
FIXTURES_DIR = Path(__file__).parent.parent / "tests" / "fixtures"

# Standard test data (same as simple.parquet)
TEST_DATA = {
    "id": [1, 2, 3, 4, 5],
    "name": ["alice", "bob", "charlie", "diana", "eve"],
    "value": [1.1, 2.2, 3.3, 4.4, 5.5],
}


def get_pyarrow_table():
    """Get standard test data as PyArrow table."""
    return pa.table({
        "id": pa.array(TEST_DATA["id"], type=pa.int64()),
        "name": pa.array(TEST_DATA["name"], type=pa.utf8()),
        "value": pa.array(TEST_DATA["value"], type=pa.float64()),
    })


def generate_arrow_ipc():
    """Generate Arrow IPC file (.arrow / .feather)."""
    print("\n[Arrow IPC]")
    table = get_pyarrow_table()

    # Arrow IPC file format (random access)
    path = FIXTURES_DIR / "simple.arrow"
    with ipc.new_file(str(path), table.schema) as writer:
        writer.write_table(table)
    print(f"  Generated: {path} ({path.stat().st_size} bytes)")

    # Also generate .feather (same format, different extension)
    path = FIXTURES_DIR / "simple.feather"
    import pyarrow.feather as feather
    feather.write_feather(table, str(path))
    print(f"  Generated: {path} ({path.stat().st_size} bytes)")

    # Arrow IPC stream format
    path = FIXTURES_DIR / "simple_stream.arrows"
    with ipc.new_stream(str(path), table.schema) as writer:
        writer.write_table(table)
    print(f"  Generated: {path} ({path.stat().st_size} bytes)")


def generate_avro():
    """Generate Avro container file (.avro)."""
    print("\n[Avro]")
    try:
        import fastavro
        from fastavro import writer as avro_writer
    except ImportError:
        print("  SKIPPED: fastavro not installed (pip install fastavro)")
        return

    schema = {
        "type": "record",
        "name": "TestRecord",
        "fields": [
            {"name": "id", "type": "long"},
            {"name": "name", "type": "string"},
            {"name": "value", "type": "double"},
        ]
    }

    records = [
        {"id": i, "name": n, "value": v}
        for i, n, v in zip(TEST_DATA["id"], TEST_DATA["name"], TEST_DATA["value"])
    ]

    # Uncompressed
    path = FIXTURES_DIR / "simple.avro"
    with open(path, "wb") as f:
        avro_writer(f, schema, records)
    print(f"  Generated: {path} ({path.stat().st_size} bytes)")

    # Deflate compression
    path = FIXTURES_DIR / "simple_deflate.avro"
    with open(path, "wb") as f:
        avro_writer(f, schema, records, codec="deflate")
    print(f"  Generated: {path} ({path.stat().st_size} bytes)")


def generate_orc():
    """Generate ORC file (.orc)."""
    print("\n[ORC]")
    try:
        import pyorc
    except ImportError:
        print("  SKIPPED: pyorc not installed (pip install pyorc)")
        return

    # ORC schema
    schema = "struct<id:bigint,name:string,value:double>"

    path = FIXTURES_DIR / "simple.orc"
    with open(path, "wb") as f:
        with pyorc.Writer(f, schema) as writer:
            for i, n, v in zip(TEST_DATA["id"], TEST_DATA["name"], TEST_DATA["value"]):
                writer.write((i, n, v))
    print(f"  Generated: {path} ({path.stat().st_size} bytes)")

    # With Snappy compression
    path = FIXTURES_DIR / "simple_snappy.orc"
    with open(path, "wb") as f:
        with pyorc.Writer(f, schema, compression=pyorc.CompressionKind.SNAPPY) as writer:
            for i, n, v in zip(TEST_DATA["id"], TEST_DATA["name"], TEST_DATA["value"]):
                writer.write((i, n, v))
    print(f"  Generated: {path} ({path.stat().st_size} bytes)")


def generate_excel():
    """Generate Excel file (.xlsx)."""
    print("\n[Excel]")
    try:
        import openpyxl
    except ImportError:
        print("  SKIPPED: openpyxl not installed (pip install openpyxl)")
        return

    path = FIXTURES_DIR / "simple.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row
    ws.append(["id", "name", "value"])

    # Data rows
    for i, n, v in zip(TEST_DATA["id"], TEST_DATA["name"], TEST_DATA["value"]):
        ws.append([i, n, v])

    wb.save(path)
    print(f"  Generated: {path} ({path.stat().st_size} bytes)")

    # Multi-sheet workbook
    path = FIXTURES_DIR / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Users"
    ws1.append(["id", "name"])
    for i, n in zip(TEST_DATA["id"], TEST_DATA["name"]):
        ws1.append([i, n])

    ws2 = wb.create_sheet("Values")
    ws2.append(["id", "value"])
    for i, v in zip(TEST_DATA["id"], TEST_DATA["value"]):
        ws2.append([i, v])

    wb.save(path)
    print(f"  Generated: {path} ({path.stat().st_size} bytes)")


def generate_delta():
    """Generate Delta Lake table (directory)."""
    print("\n[Delta Lake]")
    try:
        from deltalake import write_deltalake, DeltaTable
    except ImportError:
        print("  SKIPPED: deltalake not installed (pip install deltalake)")
        return

    table = get_pyarrow_table()

    path = FIXTURES_DIR / "simple.delta"
    if path.exists():
        shutil.rmtree(path)

    write_deltalake(str(path), table)

    # Calculate total size
    total_size = sum(f.stat().st_size for f in path.rglob("*") if f.is_file())
    print(f"  Generated: {path} ({total_size} bytes total)")

    # Show structure
    print(f"    _delta_log/: transaction logs")
    for f in sorted(path.glob("_delta_log/*.json")):
        print(f"      {f.name}")


def generate_iceberg():
    """Generate Iceberg table (directory)."""
    print("\n[Iceberg]")
    try:
        from pyiceberg.catalog import load_catalog
        from pyiceberg.schema import Schema
        from pyiceberg.types import NestedField, LongType, StringType, DoubleType
    except ImportError:
        print("  SKIPPED: pyiceberg not installed (pip install pyiceberg)")
        return

    # Iceberg requires a catalog - use in-memory SQLite
    path = FIXTURES_DIR / "simple.iceberg"
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True)

    # Create catalog config
    catalog_config = {
        "type": "sql",
        "uri": f"sqlite:///{path}/catalog.db",
        "warehouse": str(path),
    }

    try:
        catalog = load_catalog("test_catalog", **catalog_config)

        schema = Schema(
            NestedField(1, "id", LongType(), required=True),
            NestedField(2, "name", StringType(), required=True),
            NestedField(3, "value", DoubleType(), required=True),
        )

        # Create namespace and table
        catalog.create_namespace("test_ns")
        table = catalog.create_table("test_ns.simple", schema=schema)

        # Append data
        arrow_table = get_pyarrow_table()
        table.append(arrow_table)

        total_size = sum(f.stat().st_size for f in path.rglob("*") if f.is_file())
        print(f"  Generated: {path} ({total_size} bytes total)")
    except Exception as e:
        print(f"  ERROR: {e}")
        # Create minimal Iceberg-like structure for testing
        print("  Creating minimal Iceberg structure for testing...")

        metadata_dir = path / "metadata"
        metadata_dir.mkdir(parents=True, exist_ok=True)

        # Write a simple metadata JSON
        metadata = {
            "format-version": 2,
            "table-uuid": "12345678-1234-1234-1234-123456789012",
            "location": str(path),
            "schema": {
                "type": "struct",
                "schema-id": 0,
                "fields": [
                    {"id": 1, "name": "id", "required": True, "type": "long"},
                    {"id": 2, "name": "name", "required": True, "type": "string"},
                    {"id": 3, "name": "value", "required": True, "type": "double"},
                ]
            },
            "current-schema-id": 0,
            "partition-spec": [],
            "default-spec-id": 0,
            "snapshots": [],
        }

        metadata_path = metadata_dir / "v1.metadata.json"
        with open(metadata_path, "w") as f:
            json.dump(metadata, f, indent=2)

        # Write data as Parquet
        data_dir = path / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        pq.write_table(get_pyarrow_table(), data_dir / "00000-0.parquet")

        print(f"  Generated minimal structure: {path}")


def print_summary():
    """Print summary of generated fixtures."""
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)

    formats = {
        ".arrow": "Arrow IPC",
        ".feather": "Feather",
        ".arrows": "Arrow Stream",
        ".avro": "Avro",
        ".orc": "ORC",
        ".xlsx": "Excel",
        ".delta": "Delta Lake",
        ".iceberg": "Iceberg",
    }

    for ext, name in formats.items():
        if ext.startswith("."):
            files = list(FIXTURES_DIR.glob(f"*{ext}"))
            dirs = list(FIXTURES_DIR.glob(f"*{ext}"))
        else:
            files = []
            dirs = [d for d in FIXTURES_DIR.iterdir() if d.is_dir() and d.name.endswith(ext)]

        items = files + [d for d in dirs if d.is_dir()]
        if items:
            print(f"\n{name}:")
            for item in sorted(items):
                if item.is_file():
                    print(f"  {item.name}: {item.stat().st_size} bytes")
                else:
                    total = sum(f.stat().st_size for f in item.rglob("*") if f.is_file())
                    print(f"  {item.name}/: {total} bytes")


def main():
    print("=" * 60)
    print("LanceQL Test Fixture Generator")
    print("Generating fixtures for all supported formats")
    print("=" * 60)

    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)

    generate_arrow_ipc()
    generate_avro()
    generate_orc()
    generate_excel()
    generate_delta()
    generate_iceberg()

    print_summary()


if __name__ == "__main__":
    main()
